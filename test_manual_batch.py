#!/usr/bin/env python3
"""
Manual test for batch processing - simulates the exact steps
"""
import os
import glob
import pandas as pd
from openpyxl import load_workbook

def test_batch_processing_manual():
    """Manually test the batch processing logic"""
    print("=== MANUAL BATCH PROCESSING TEST ===")
    print()
    
    # Set up paths
    source_folder = 'tests/test_data/Input'
    target_folder = 'tests/test_data/Output'
    pattern = 'target_enrollment*.xlsx'
    
    print("1. Setting up test...")
    print(f"   Source folder: {source_folder}")
    print(f"   Target folder: {target_folder}")
    print(f"   Pattern: {pattern}")
    print()
    
    # Step 1: Find source files
    source_file_pattern = os.path.join(source_folder, "*.xlsx")
    source_files = glob.glob(source_file_pattern)
    
    print(f"2. Found {len(source_files)} source files:")
    for f in source_files:
        print(f"   - {os.path.basename(f)}")
    
    if not source_files:
        print("ERROR: No source files found!")
        return False
    
    # Use first source file
    source_file = source_files[0]
    print(f"   Using: {os.path.basename(source_file)}")
    print()
    
    # Step 2: Find target files
    target_file_pattern = os.path.join(target_folder, pattern)
    target_files = glob.glob(target_file_pattern)
    
    print(f"3. Found {len(target_files)} target files:")
    for f in target_files:
        print(f"   - {os.path.basename(f)}")
    
    if not target_files:
        print("ERROR: No target files found!")
        return False
    
    print()
    
    # Step 3: Read source data
    print("4. Reading source data...")
    source_df = pd.read_excel(source_file)
    print(f"   Shape: {source_df.shape}")
    print(f"   Columns: {list(source_df.columns)}")
    
    # Find CRN and Actual columns
    crn_options = ["'CRN'", "CRN"]
    actual_options = ["'Actual'", "Actual"]
    
    crn_col = next((col for col in crn_options if col in source_df.columns), None)
    actual_col = next((col for col in actual_options if col in source_df.columns), None)
    
    print(f"   CRN column: {crn_col}")
    print(f"   Actual column: {actual_col}")
    
    if not crn_col or not actual_col:
        print("ERROR: Required columns not found!")
        return False
    
    # Create enrollment map
    enrollment_map = {}
    for _, row in source_df.iterrows():
        crn = str(row[crn_col]).strip()
        if crn and pd.notna(row[actual_col]):
            enrollment_map[crn] = int(row[actual_col])
    
    print(f"   Enrollment map: {enrollment_map}")
    print()
    
    # Step 4: Process each target file
    print("5. Processing target files...")
    total_updates = 0
    
    for target_file in target_files:
        print(f"   Processing: {os.path.basename(target_file)}")
        
        # Load target workbook
        target_wb = load_workbook(target_file)
        target_ws = target_wb.active
        
        # Find header row
        header_row = [str(cell.value) if cell.value is not None else '' for cell in target_ws[1]]
        print(f"     Headers: {header_row}")
        
        # Find column indices
        target_crn_idx = None
        target_enroll_idx = None
        
        for i, col_name in enumerate(header_row):
            if col_name == 'CRN':
                target_crn_idx = i
            elif col_name == 'Current Enroll':
                target_enroll_idx = i
        
        print(f"     CRN column index: {target_crn_idx}")
        print(f"     Current Enroll column index: {target_enroll_idx}")
        
        if target_crn_idx is None or target_enroll_idx is None:
            print("     ERROR: Required columns not found!")
            continue
        
        # Update rows
        file_updates = 0
        for row_idx, row in enumerate(target_ws.iter_rows(min_row=2), 2):
            target_crn = str(row[target_crn_idx].value).strip() if row[target_crn_idx].value is not None else ""
            
            if target_crn in enrollment_map:
                old_value = row[target_enroll_idx].value
                new_value = enrollment_map[target_crn]
                
                if old_value != new_value:
                    print(f"     Row {row_idx}: CRN {target_crn} - {old_value} → {new_value}")
                    row[target_enroll_idx].value = new_value
                    file_updates += 1
        
        # Save the file
        target_wb.save(target_file)
        print(f"     Updates made: {file_updates}")
        total_updates += file_updates
        print()
    
    print(f"6. Total updates made: {total_updates}")
    
    if total_updates > 0:
        print("✓ BATCH PROCESSING TEST PASSED!")
        return True
    else:
        print("✗ BATCH PROCESSING TEST FAILED - No updates made")
        return False

if __name__ == "__main__":
    success = test_batch_processing_manual()
    exit(0 if success else 1)
