#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Simplified Enrollment Updater
Just matches CRN and updates Current Enroll values from Actual column
"""

import tkinter as tk
import os
import glob
from tkinter import filedialog, messagebox, scrolledtext
from openpyxl import load_workbook
import pandas as pd

class ExcelDataReplacer:
    def __init__(self, master):
        self.master = master
        master.title("Soc & Crim Enrollment Updater")
        master.geometry('600x400')

        self.label = tk.Label(master, text="Soc & Crim Enrollment Updater", font=('Times New Roman', 14))
        self.label.pack(pady=10)

        # Frame for processing mode
        self.mode_frame = tk.Frame(master)
        self.mode_frame.pack(pady=5)
        
        # Processing mode selection
        self.process_mode = tk.StringVar(value="single")
        self.radio_single = tk.Radiobutton(self.mode_frame, text="Process Single Files", 
                                          variable=self.process_mode, value="single",
                                          command=self.toggle_mode)
        self.radio_batch = tk.Radiobutton(self.mode_frame, text="Process Folder", 
                                         variable=self.process_mode, value="folder",
                                         command=self.toggle_mode)
        self.radio_single.pack(side=tk.LEFT, padx=5)
        self.radio_batch.pack(side=tk.LEFT, padx=5)

        # Frame for source selection
        self.source_frame = tk.Frame(master)
        self.source_frame.pack(pady=5, fill=tk.X, padx=20)
        
        self.source_label = tk.Label(self.source_frame, text="Source File:", width=10, anchor="w")
        self.source_label.pack(side=tk.LEFT)
        
        self.source_entry = tk.Entry(self.source_frame)
        self.source_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.btn_source = tk.Button(self.source_frame, text="Browse", command=self.select_source)
        self.btn_source.pack(side=tk.RIGHT)

        # Frame for target selection
        self.target_frame = tk.Frame(master)
        self.target_frame.pack(pady=5, fill=tk.X, padx=20)
        
        self.target_label = tk.Label(self.target_frame, text="Target File:", width=10, anchor="w")
        self.target_label.pack(side=tk.LEFT)
        
        self.target_entry = tk.Entry(self.target_frame)
        self.target_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        self.btn_target = tk.Button(self.target_frame, text="Browse", command=self.select_target)
        self.btn_target.pack(side=tk.RIGHT)

        # Frame for file pattern when in folder mode
        self.pattern_frame = tk.Frame(master)
        self.pattern_frame.pack(pady=5, fill=tk.X, padx=20)
        
        self.pattern_label = tk.Label(self.pattern_frame, text="Pattern:", width=10, anchor="w")
        self.pattern_label.pack(side=tk.LEFT)
        
        self.pattern_entry = tk.Entry(self.pattern_frame)
        self.pattern_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.pattern_entry.insert(0, "*.xlsx")
        
        # Initially hide pattern frame for single file mode
        self.pattern_frame.pack_forget()

        # Frame for buttons 
        self.button_frame = tk.Frame(master)
        self.button_frame.pack(pady=5)
        
        # Execute button
        self.btn_execute = tk.Button(self.button_frame, text="Update Enrollment", command=self.replace_data, width=15)
        self.btn_execute.pack(side=tk.LEFT, padx=5)

        # Add a view updates button
        self.btn_view_updates = tk.Button(self.button_frame, text="View Last Updates", command=self.show_updates_window, width=15)
        self.btn_view_updates.pack(side=tk.LEFT, padx=5)
        
        # Add debug info
        self.debug_var = tk.BooleanVar(value=True)  # Default to True for debugging
        self.debug_checkbox = tk.Checkbutton(master, text="Debug Mode (verbose logging)", 
                                            variable=self.debug_var)
        self.debug_checkbox.pack(pady=5)

        # Status label
        self.status_label = tk.Label(master, text="", font=('Times New Roman', 10))
        self.status_label.pack(pady=5)

        self.source_file = None
        self.source_folder = None
        self.target_file = None
        self.target_folder = None
        
        # Store update history
        self.updates = []
        self.all_update_details = []  # For batch processing

    def toggle_mode(self):
        """Toggle between single file and folder modes"""
        if self.process_mode.get() == "single":
            self.source_label.config(text="Source File:")
            self.target_label.config(text="Target File:")
            self.pattern_frame.pack_forget()
            self.source_entry.delete(0, tk.END)
            self.target_entry.delete(0, tk.END)
        else:
            self.source_label.config(text="Source Dir:")
            self.target_label.config(text="Target Dir:")
            self.pattern_frame.pack(pady=5, fill=tk.X, padx=20, after=self.target_frame)
            self.source_entry.delete(0, tk.END)
            self.target_entry.delete(0, tk.END)

    def select_source(self):
        """Select source file or folder based on mode"""
        if self.process_mode.get() == "single":
            self.source_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
            if self.source_file:
                self.source_entry.delete(0, tk.END)
                self.source_entry.insert(0, self.source_file)
        else:
            self.source_folder = filedialog.askdirectory()
            if self.source_folder:
                self.source_entry.delete(0, tk.END)
                self.source_entry.insert(0, self.source_folder)

    def select_target(self):
        """Select target file or folder based on mode"""
        if self.process_mode.get() == "single":
            self.target_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
            if self.target_file:
                self.target_entry.delete(0, tk.END)
                self.target_entry.insert(0, self.target_file)
        else:
            self.target_folder = filedialog.askdirectory()
            if self.target_folder:
                self.target_entry.delete(0, tk.END)
                self.target_entry.insert(0, self.target_folder)

    def update_status(self, message):
        """Update status message"""
        self.status_label.config(text=message)
        self.master.update_idletasks()

    def replace_data(self):
        """Process single file or batch based on mode"""
        if self.process_mode.get() == "single":
            self.replace_data_single()
        else:
            self.replace_data_batch()

    def replace_data_single(self):
        """Process a single source and target file"""
        source_file = self.source_entry.get()
        target_file = self.target_entry.get()
        
        if not source_file or not target_file:
            messagebox.showerror("Error", "Please select both source and target Excel files.")
            return

        try:
            self.update_status(f"Processing: {os.path.basename(source_file)} -> {os.path.basename(target_file)}")
            updated, update_details = self.process_files(source_file, target_file)
            
            if updated > 0:
                messagebox.showinfo("Success", f"Enrollment data updated successfully!\n\n"
                                  f"• {updated} rows were updated\n"
                                  f"• Click 'View Last Updates' to see details")
                self.updates = update_details
                self.btn_view_updates.config(state=tk.NORMAL)
            else:
                messagebox.showinfo("Info", "No matching rows found for update.")
                
            self.update_status("Done")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")
            self.update_status("Error occurred")

    def replace_data_batch(self):
        """Process all files in source folder matching pattern"""
        source_folder = self.source_entry.get()
        target_folder = self.target_entry.get()
        pattern = self.pattern_entry.get()
        
        if not source_folder or not target_folder or not pattern:
            messagebox.showerror("Error", "Please select both source and target folders and specify a file pattern.")
            return

        try:
            file_pattern = os.path.join(source_folder, pattern)
            source_files = glob.glob(file_pattern)
            
            if not source_files:
                messagebox.showinfo("Info", f"No files matching '{pattern}' found in the source folder.")
                return
                
            total_updated = 0
            files_processed = 0
            self.all_update_details = []  # Reset for batch processing
            
            for source_file in source_files:
                file_name = os.path.basename(source_file)
                target_file = os.path.join(target_folder, file_name)
                
                # Skip if target file doesn't exist
                if not os.path.exists(target_file):
                    self.update_status(f"Skipping {file_name}: Target file doesn't exist")
                    continue
                    
                self.update_status(f"Processing: {file_name}")
                try:
                    rows_updated, update_details = self.process_files(source_file, target_file)
                    
                    if rows_updated > 0:
                        file_updates = {
                            'file': file_name,
                            'details': update_details
                        }
                        self.all_update_details.append(file_updates)
                        
                    total_updated += rows_updated
                    files_processed += 1
                    self.update_status(f"Updated {rows_updated} rows in {file_name}")
                except Exception as e:
                    self.update_status(f"Error processing {file_name}: {str(e)}")
            
            if total_updated > 0:
                messagebox.showinfo("Batch Complete", 
                                  f"Batch processing successful!\n\n"
                                  f"• Processed {files_processed} files\n"
                                  f"• Total of {total_updated} rows updated\n"
                                  f"• Click 'View Last Updates' to see details")
                self.btn_view_updates.config(state=tk.NORMAL)
            else:
                messagebox.showinfo("Batch Complete", 
                                  f"Batch processing complete\n\n"
                                  f"• Processed {files_processed} files\n"
                                  f"• No matching rows found for update")
                
            self.update_status(f"Batch processing complete. {files_processed} files processed.")
                
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during batch processing: {e}")
            self.update_status("Error occurred")

    def show_updates_window(self):
        """Display a window showing the details of what was updated"""
        update_window = tk.Toplevel(self.master)
        update_window.title("Enrollment Update Details")
        update_window.geometry("800x600")
        
        # Create a frame for the heading
        heading_frame = tk.Frame(update_window, bg="#f0f0f0")
        heading_frame.pack(fill=tk.X, padx=10, pady=5)
        
        heading_label = tk.Label(heading_frame, text="=== ENROLLMENT UPDATE DETAILS ===", 
                              font=("Arial", 14, "bold"), bg="#f0f0f0")
        heading_label.pack(pady=10)
        
        # Create an info label to show summary
        if self.process_mode.get() == "single":
            update_count = len(self.updates) if hasattr(self, 'updates') and self.updates else 0
            info_text = f"Showing {update_count} updates"
        else:
            file_count = len(self.all_update_details) if hasattr(self, 'all_update_details') and self.all_update_details else 0
            update_count = sum(len(file_update['details']) for file_update in self.all_update_details) if file_count > 0 else 0
            info_text = f"Showing {update_count} updates across {file_count} files"
            
        info_label = tk.Label(heading_frame, text=info_text, font=("Arial", 10), bg="#f0f0f0")
        info_label.pack(pady=5)
        
        # Create a frame for the text area
        text_frame = tk.Frame(update_window)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Create a scrolled text widget to display the updates
        text_area = scrolledtext.ScrolledText(text_frame, width=80, height=25, font=("Consolas", 10))
        text_area.pack(fill=tk.BOTH, expand=True)
        
        # Configure text tags for formatting
        text_area.tag_configure("heading", font=("Arial", 12, "bold"), foreground="navy")
        text_area.tag_configure("subheading", font=("Arial", 11, "bold"), foreground="blue")
        text_area.tag_configure("bold", font=("Arial", 10, "bold"))
        text_area.tag_configure("field", font=("Arial", 9, "bold"))
        text_area.tag_configure("changed", foreground="green")
        text_area.tag_configure("file_heading", font=("Arial", 11, "bold"), foreground="blue")
        
        # Add a heading
        text_area.insert(tk.END, "=== ENROLLMENT UPDATE DETAILS ===\n\n", "heading")
        
        if (self.process_mode.get() == "single" and hasattr(self, 'updates') and self.updates) or \
           (self.process_mode.get() != "single" and hasattr(self, 'all_update_details') and self.all_update_details):
            if self.process_mode.get() == "single":
                self._show_single_file_updates(text_area)
            else:
                self._show_batch_updates(text_area)
        else:
            text_area.insert(tk.END, "No updates to display.\n\nPerform an enrollment update first to see details here.")
        
        # Make the text widget read-only
        text_area.config(state=tk.DISABLED)
        
        # Create a frame for buttons
        button_frame = tk.Frame(update_window)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Add a close button
        close_btn = tk.Button(button_frame, text="Close", command=update_window.destroy)
        close_btn.pack(side=tk.RIGHT, padx=5)

    def _show_single_file_updates(self, text_area):
        """Show updates for a single file with improved formatting"""
        if not self.updates:
            text_area.insert(tk.END, "No updates to display.")
            return
        
        # Show statistics
        total_rows = len(self.updates)
        total_fields = sum(len(update['fields']) for update in self.updates)
        
        text_area.insert(tk.END, f"Summary:\n", "subheading")
        text_area.insert(tk.END, f"- Total rows updated: {total_rows}\n")
        text_area.insert(tk.END, f"- Total fields changed: {total_fields}\n\n")
        text_area.insert(tk.END, "Details:\n", "subheading")
        text_area.insert(tk.END, "-" * 60 + "\n\n")
        
        # Show detailed updates
        for update in self.updates:
            row_num = update['row']
            crn = update['crn']
            
            text_area.insert(tk.END, f"Row {row_num}: ", "bold")
            text_area.insert(tk.END, f"CRN={crn}\n")
            
            if not update['fields']:
                text_area.insert(tk.END, "  No fields were changed\n")
            else:
                text_area.insert(tk.END, "  Fields updated:\n")
            
                for field, values in update['fields'].items():
                    old_val = values.get('old', 'N/A')
                    new_val = values.get('new', 'N/A')
                    
                    # Skip fields that didn't actually change
                    if str(old_val).strip() == str(new_val).strip():
                        continue
                        
                    text_area.insert(tk.END, f"    - {field}: ", "field")
                    
                    # Format old and new values properly
                    old_val_str = str(old_val) if old_val is not None else "(empty)"
                    new_val_str = str(new_val) if new_val is not None else "(empty)"
                    
                    text_area.insert(tk.END, f"{old_val_str} → ", "")
                    text_area.insert(tk.END, f"{new_val_str}\n", "changed")
            
            text_area.insert(tk.END, "\n" + "-" * 60 + "\n\n")
    
    def _show_batch_updates(self, text_area):
        """Show updates for batch processing with improved formatting"""
        if not self.all_update_details:
            text_area.insert(tk.END, "No updates to display.")
            return
        
        # Calculate summary statistics
        total_files = len(self.all_update_details)
        total_rows = sum(len(file_update['details']) for file_update in self.all_update_details)
        total_fields_changed = 0
        
        for file_update in self.all_update_details:
            for update in file_update['details']:
                total_fields_changed += sum(1 for field, values in update['fields'].items() 
                                         if str(values.get('old', '')) != str(values.get('new', '')))
        
        # Display summary
        text_area.insert(tk.END, f"Batch Summary:\n", "subheading")
        text_area.insert(tk.END, f"- Files processed: {total_files}\n")
        text_area.insert(tk.END, f"- Total rows updated: {total_rows}\n")
        text_area.insert(tk.END, f"- Total fields changed: {total_fields_changed}\n\n")
        text_area.insert(tk.END, "File Details:\n", "subheading")
        text_area.insert(tk.END, "=" * 70 + "\n\n")
            
        # Show per-file updates
        for file_update in self.all_update_details:
            file_name = file_update['file']
            details = file_update['details']
            
            if not details:
                continue
            
            # Count updates for this file
            file_rows = len(details)
            file_fields = sum(len(update['fields']) for update in details)
                
            text_area.insert(tk.END, f"File: {file_name}\n", "file_heading")
            text_area.insert(tk.END, "=" * len(f"File: {file_name}") + "\n")
            text_area.insert(tk.END, f"- Rows updated: {file_rows}, Fields changed: {file_fields}\n\n")
            
            for update in details:
                row_num = update['row']
                crn = update['crn']
                
                text_area.insert(tk.END, f"Row {row_num}: ", "bold")
                text_area.insert(tk.END, f"CRN={crn}\n")
                
                if not update['fields']:
                    text_area.insert(tk.END, "  No fields were changed\n")
                else:
                    text_area.insert(tk.END, "  Fields updated:\n")
                
                    for field, values in update['fields'].items():
                        old_val = values.get('old', 'N/A')
                        new_val = values.get('new', 'N/A')
                        
                        # Skip fields that didn't actually change
                        if str(old_val).strip() == str(new_val).strip():
                            continue
                            
                        text_area.insert(tk.END, f"    - {field}: ", "field")
                        
                        # Format old and new values properly
                        old_val_str = str(old_val) if old_val is not None else "(empty)"
                        new_val_str = str(new_val) if new_val is not None else "(empty)"
                        
                        text_area.insert(tk.END, f"{old_val_str} → ", "")
                        text_area.insert(tk.END, f"{new_val_str}\n", "changed")
                
                text_area.insert(tk.END, "\n" + "-"*40 + "\n")
            
            text_area.insert(tk.END, "\n" + "="*70 + "\n\n")

    def find_header_row(self, ws, max_rows_to_check=50):
        """Find the header row in the target worksheet"""
        for row_idx in range(1, min(max_rows_to_check + 1, ws.max_row + 1)):
            row_values = [str(cell.value).lower() if cell.value is not None else '' for cell in ws[row_idx]]
            
            # Look for 'crn' and 'current enroll' in the same row
            if 'crn' in row_values and any('enroll' in val for val in row_values):
                # Create header map
                header_row = [str(cell.value) if cell.value is not None else '' for cell in ws[row_idx]]
                header_map = {header: i for i, header in enumerate(header_row) if header}
                if self.debug_var.get():
                    print(f"Found header row at {row_idx}")
                return row_idx, header_row, header_map
                
        # Default to first row if no header found
        first_row_values = [str(cell.value) if cell.value is not None else '' for cell in ws[1]]
        header_map = {header: i for i, header in enumerate(first_row_values) if header}
        return 1, first_row_values, header_map

    def process_files(self, source_file, target_file):
        """SIMPLIFIED version that just matches CRN and updates enrollment"""
        try:
            # Read source data
            source_df = pd.read_excel(source_file)
            if self.debug_var.get():
                print(f"Source file loaded: {source_file}")
                print(f"Source dataframe shape: {source_df.shape}")
                print(f"Source columns: {list(source_df.columns)}")
            
            # Determine source columns - try both quoted and unquoted versions
            crn_options = ["'CRN'", "CRN"]  # Try both quoted and unquoted
            actual_options = ["'Actual'", "Actual"]  # Try both quoted and unquoted
            
            # Find the correct column names
            crn_col = next((col for col in crn_options if col in source_df.columns), None)
            actual_col = next((col for col in actual_options if col in source_df.columns), None)
            
            # Check if columns exist
            if crn_col is None:
                raise ValueError(f"CRN column not found in source file. Available columns: {list(source_df.columns)}")
            if actual_col is None:
                raise ValueError(f"Actual column not found in source file. Available columns: {list(source_df.columns)}")
            
            # Load target file
            target_wb = load_workbook(target_file)
            target_ws = target_wb.active
            
            # Find header row
            header_row_idx, header, header_map = self.find_header_row(target_ws)
            if self.debug_var.get():
                print(f"Target file header at row {header_row_idx}")
                print(f"Target columns: {header}")
            
            # Determine target columns
            target_crn_idx = None
            target_enroll_idx = None
            
            # Find CRN column
            for i, col_name in enumerate(header):
                if col_name == 'CRN':
                    target_crn_idx = i
                    break
            
            # Find Current Enroll column
            for i, col_name in enumerate(header):
                if col_name == 'Current Enroll':
                    target_enroll_idx = i
                    break
            
            if target_crn_idx is None:
                raise ValueError("Column 'CRN' not found in target file")
            if target_enroll_idx is None:
                raise ValueError("Column 'Current Enroll' not found in target file")
            
            if self.debug_var.get():
                print(f"Target CRN column at index {target_crn_idx}")
                print(f"Target Current Enroll column at index {target_enroll_idx}")
            
            # Process the data
            updated = 0
            update_details = []
            
            # Create mapping of CRNs to enrollment values from source
            enrollment_map = {}
            for _, row in source_df.iterrows():
                crn = str(row[crn_col]).strip()
                if crn and pd.notna(row[actual_col]):
                    enrollment_map[crn] = int(row[actual_col])
            
            if self.debug_var.get():
                print(f"Source CRN to enrollment map created with {len(enrollment_map)} entries")
                # Show first 5 entries for debugging
                print("First 5 entries in map:")
                for i, (crn, enroll) in enumerate(list(enrollment_map.items())[:5]):
                    print(f"  {crn}: {enroll}")
            
            # Update each row in target
            for row_idx, row in enumerate(target_ws.iter_rows(min_row=header_row_idx+1), header_row_idx+1):
                # Skip any rows that might be headers
                if row[0].value is not None and str(row[0].value).strip() == "Room #":
                    continue
                
                # Get the CRN
                target_crn = str(row[target_crn_idx].value).strip() if row[target_crn_idx].value is not None else ""
                if not target_crn:
                    continue
                
                # Convert numeric CRNs to clean strings
                if target_crn and target_crn.replace('.', '', 1).isdigit():
                    if float(target_crn) == int(float(target_crn)):
                        target_crn = str(int(float(target_crn)))
                
                # Check if it's in our map
                if target_crn in enrollment_map:
                    # Get the old value
                    old_value = row[target_enroll_idx].value
                    new_value = enrollment_map[target_crn]
                    
                    if self.debug_var.get():
                        print(f"Match found - CRN {target_crn}: updating {old_value} → {new_value}")
                    
                    # Create an update record
                    row_update = {
                        'row': row_idx,
                        'crn': target_crn,
                        'fields': {
                            'Current Enroll': {
                                'old': old_value,
                                'new': new_value
                            }
                        }
                    }
                    
                    # Update the cell
                    row[target_enroll_idx].value = new_value
                    updated += 1
                    update_details.append(row_update)
            
            # Save the target file
            target_wb.save(target_file)
            
            if self.debug_var.get():
                print(f"Updated {updated} rows in target file")
            
            return updated, update_details
            
        except Exception as e:
            if self.debug_var.get():
                print(f"Error processing files: {str(e)}")
            raise e

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelDataReplacer(root)
    root.mainloop()
